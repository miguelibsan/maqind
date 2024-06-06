import openpyxl

# Leer el primer archivo
file1_path = '/Users/miguelibarra/Documentos/maqind/Python/alv.xlsx'
wb1 = openpyxl.load_workbook(file1_path)
ws1 = wb1.active

# Leer el segundo archivo
file2_path = '/Users/miguelibarra/Documentos/maqind/Python/maq.xlsx'
wb2 = openpyxl.load_workbook(file2_path)
ws2 = wb2.active

# Eliminar la primera fila del segundo archivo
ws2.delete_rows(1)
ws1.delete_rows(1)


# Copiar los datos del segundo archivo al final del primer archivo
for row in ws2.iter_rows(min_row=1, values_only=True):
    ws1.append(row)

# Especificar las columnas a eliminar (basado en el índice, por ejemplo, B, C, E, F, G, H, I, J, K, L, M, N, O, P, Q, R, S, T, U)
columns_to_delete = ['B', 'C', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U']
for col in sorted(columns_to_delete, reverse=True):
    ws1.delete_cols(openpyxl.utils.column_index_from_string(col))

# Guardar el resultado en un nuevo archivo
output_path = '/Users/miguelibarra/Documentos/maqind/Python/smo.xlsx'
wb1.save(output_path)

print("Los datos se han combinado y guardado en 'smo.xlsx' con las columnas especificadas eliminadas.")