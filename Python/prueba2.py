import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Leer los archivos de Excel
file1_path = '/Users/miguelibarra/Documentos/maqind/Python/smo.xlsx'
file2_path = '/Users/miguelibarra/Documentos/maqind/Python/planeado.xlsx'

df1 = pd.read_excel(file1_path)
df2 = pd.read_excel(file2_path)

# Suponiendo que la columna a comparar es 'A', cambiar el nombre según corresponda
col_name = 'A'

# Encontrar valores comunes y no comunes
common_values = set(df1[col_name]).intersection(set(df2[col_name]))
only_in_df1 = set(df1[col_name]).difference(set(df2[col_name]))
only_in_df2 = set(df2[col_name]).difference(set(df1[col_name]))

# Cargar el archivo planeado.xlsx usando openpyxl para colorear las celdas
wb = load_workbook(file2_path)
ws = wb.active

# Definir los colores de relleno
green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

# Colorear celdas según las condiciones
for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
    for cell in row:
        if cell.value in common_values:
            cell.fill = green_fill
        elif cell.value in only_in_df2:
            cell.fill = orange_fill

# Guardar los cambios en el archivo planeado.xlsx
wb.save(file2_path)

print("Las celdas comunes se han coloreado en verde claro y las no comunes en naranja en el archivo 'planeado.xlsx'.")